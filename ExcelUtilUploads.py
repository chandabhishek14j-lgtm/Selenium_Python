from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl
import os
import time

#going to the url and downloading the Excel
driver = webdriver.Chrome()
wait = WebDriverWait(driver,5)
url = "https://rahulshettyacademy.com/upload-download-test/"
driver.get(url)
driver.implicitly_wait(5)
button = wait.until(EC.presence_of_element_located((By.ID, "downloadButton")))
button.click()
file_path = "C:\\Users\\AVITA\\Downloads\\download.xlsx"

# working with Excel checking if the file download is complete and file exists in local
while True:
    if os.path.exists(file_path) and not os.path.exists(file_path+".crdownload"):
        break
    time.sleep(1)
excel_dict = {}
fruit_name = "Apple"
def update_excel(path, row_item, col_name, new_val ):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active

    # traversing through all cells to find the desired cell to update


    for i in range(1, sheet.max_column + 1):
        if sheet.cell(row=1, column=i).value == col_name:
            excel_dict["column"] = i

    for i in range(1, sheet.max_row + 1):
        for j in range(1, sheet.max_column + 1):
            if sheet.cell(row=i, column=j).value == row_item:
                excel_dict["row"] = i

    # we have captured the row and col in dictionary we can use them to edit value
    sheet.cell(row=excel_dict["row"], column=excel_dict["column"]).value = new_val

    # saving the updated file
    workbook.save(file_path)

#uploading the file back to browser
#for uploading any file from local to browser...
#...DOM needs to have a type = file attribute as...
#...selenium on its own can not interact with window
update_excel(file_path, fruit_name, "price", "555")
fileElement = driver.find_element(By.CSS_SELECTOR, "input[type='file']")
"""After this we need to send path of the file in send keys method """
fileElement.send_keys(file_path)
toast = wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, ".Toastify__toast")))
UpdatedText = toast.text
print("Toast found")
print()

SuccessText = "Updated Excel Data Successfully."
assert SuccessText == UpdatedText