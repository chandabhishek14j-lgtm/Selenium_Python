import openpyxl

book = openpyxl.load_workbook("C:\\Users\\AVITA\\Desktop\\SeleniumPython\\Excel\\Excel1.xlsx")

sheet = book.active #selecting active sheet from book

cell1 = sheet.cell(row=1,column=1)
print(cell1.value)
cell2 = sheet.cell(row=1,column=2)
print(cell2.value)
cell3 = sheet.cell(row=1,column=3)
print(cell3.value)
cell4 = sheet.cell(row=1,column=4)
print(cell4.value)
cell5 = sheet.cell(row=1,column=5)
print(cell5.value)
cell6 = sheet.cell(row=2,column=1)
print(cell6.value)
cell7 = sheet.cell(row=3,column=1)
print(cell7.value)

# now to write anything on Excel sheet reverse the sheet format
sheet.cell(row=2, column=2).value = "Abhishek"
cell8 = sheet.cell(row=2, column=2)
print(cell8.value)

#shortcut to print value of any cell using name box of the cell
shortCutCell = sheet['A1'].value
print(shortCutCell)

# how to find sheet level no. of rows and no. of columns
rowCount = sheet.max_row
columnCount = sheet.max_column
print(rowCount)
print(columnCount)
print("*********************")

# how to read whole data of Excel
for item in range(1, rowCount+1):#rowcount+1 because last number is excluded in range for python
    for j in range(1, columnCount+1):
        print(sheet.cell(row=item, column=j).value)

print("*********************")

#using conditional statements to get specific row data
for item in range(1, rowCount+1):
    if sheet.cell(row=item, column=1).value == "TestCase2":
        for j in range(1, columnCount+1):
            print(sheet.cell(row=item, column=j).value)
    else:
        continue
print("*********--Dictionary--************")
#storing all data into a dictionary
abc = {}# initializing empty dictionary
for item in range(1, rowCount+1):
    if sheet.cell(row=item, column=1).value == "TestCase2":
        for j in range(1, columnCount+1):
            abc[sheet.cell(row=1, column=j).value] = sheet.cell(row=item, column=j).value
    else:
        continue
print(abc)